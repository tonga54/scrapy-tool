import time, csv, urllib.request, xmltodict, json
from openpyxl import load_workbook

class Tool:

    @staticmethod
    def getAlexaRank(searchedData, min_popularity, max_popularity):
        try:
            response = dict()
            for dt in searchedData:
                xml = urllib.request.urlopen('http://data.alexa.com/data?cli=10&dat=s&url={}'.format(dt)).read()
                result= xmltodict.parse(xml)
                data = json.dumps(result).replace("@","")
                data_tojson = json.loads(data)
                if type(data_tojson["ALEXA"]["SD"]) is list:
                    if "COUNTRY" in data_tojson["ALEXA"]["SD"][1] and "POPULARITY" in data_tojson["ALEXA"]["SD"][1]:
                        countryName = data_tojson["ALEXA"]["SD"][1]["COUNTRY"]["NAME"]
                        countryRank = data_tojson["ALEXA"]["SD"][1]["COUNTRY"]["RANK"]
                        popularity = int(data_tojson["ALEXA"]["SD"][1]["POPULARITY"]["TEXT"])
                        if popularity >= min_popularity and popularity <= max_popularity:
                            response[dt] = searchedData[dt]
                            response[dt]['alexa_rank'] = {
                                "popularity": popularity,
                                "countryName": countryName,
                                "countryRank": countryRank
                            }

            return response
        except Exception as e:
            return e

    # @staticmethod
    # def createCsv():
    #     with open('data.csv', 'w', newline='') as file:
    #         writer = csv.writer(file)
    #         writer.writerow(["Url", "Keywords Rank", "Popularity", "Country Name", "Country Rank"])

    # @staticmethod
    # def writeInCsv(url, data):
    #     with open('data.csv', 'a') as file:
    #         writer = csv.writer(file)
    #         writer.writerow([url, data['keywords_rank'], data['alexa_rank']['popularity'], data['alexa_rank']['countryName'], data['alexa_rank']['countryRank']])

    @staticmethod
    def readFile(file):
        wb = load_workbook(file)
        sheet_obj = wb.active 
        m_row = sheet_obj.max_row 
        
        webs = []
        for i in range(1, m_row + 1): 
            cell_obj = sheet_obj.cell(row = i, column = 2)
            if cell_obj.value is not None:
                webs.append(cell_obj.value)

        return webs

    @staticmethod
    def analyzeSite(text, keywords):
        try:
            data = dict()
            data['keywords_rank'] = 0
            for kw in keywords:
                if kw in text:
                    data['keywords_rank'] += 1

            if data['keywords_rank'] != 0:
                data['keywords_rank'] = (data['keywords_rank'] * 100) / len(keywords) 
            return data
        except Exception as e:
            return e